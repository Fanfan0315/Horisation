from io import BytesIO
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def detect_separator(binary: bytes, encodings=('utf-8', 'utf-8-sig', 'latin1', 'gbk')):
    """
    自动检测 CSV 文件的分隔符和编码
    - 默认尝试多种常见编码
    - 支持常见分隔符：逗号、分号、制表符、竖线
    """
    for enc in encodings:
        try:
            text = binary.decode(enc)
            sample = text[:10000]  # 取前一段文本作为样本
            # 用 Sniffer 猜分隔符
            dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t', '|'])
            print(f"[DEBUG] 检测到分隔符: {repr(dialect.delimiter)}, 编码: {enc}")
            return dialect.delimiter, enc
        except Exception:
            continue

    # 如果都失败，默认逗号 + utf-8
    print("[DEBUG] 自动检测失败，使用默认分隔符 ',' 和编码 'utf-8'")
    return ',', 'utf-8'

def _to_df(binary: bytes) -> pd.DataFrame:
    """
    将二进制CSV读成DataFrame。
    - 自动推断编码（常见UTF-8/GBK情况pandas会处理；若有特殊编码可扩展）
    """
    # 你也可以在 read_csv 里加参数，如 sep=';', encoding='utf-8', dtype=str 等
    sep, encoding = detect_separator(binary)
    bio = BytesIO(binary)
    df = pd.read_csv(bio, sep=sep, encoding=encoding)
    df.attrs["sep"] = sep
    df.attrs["encoding"] = encoding
    return df

def read_csv_preview(binary: bytes, n: int = 5):
    """返回前 n 行的预览记录（list[dict]）与列名"""
    df = _to_df(binary)
    head = df.head(n)
    return {
        'columns': list(head.columns),
        'rows': head.to_dict(orient='records')
    }

def summarize_csv(binary: bytes):
    """返回整体概要信息：行/列、字段类型、缺失统计等"""
    df = _to_df(binary)

    # 字段类型（pandas dtype -> 简单字符串）
    dtypes = {col: str(dt) for col, dt in df.dtypes.items()}

    # 缺失值计数与占比
    na_count = df.isna().sum().to_dict()
    total_rows = len(df)
    na_ratio = {k: (v / total_rows if total_rows else 0.0) for k, v in na_count.items()}

    summary = {
        'rows': int(total_rows),
        'cols': int(df.shape[1]),
        'columns': list(df.columns),
        'dtypes': dtypes,
        'na_count': na_count,
        'na_ratio': na_ratio
    }
    return summary

def clean_column_name(df: pd.DataFrame, cols: list[str] | None = None) -> pd.DataFrame:
    target_cols = df.columns if cols is None else cols
    new_cols = []
    for col in df.columns:
        if col in target_cols:
            new_col = col.strip()
            new_col = new_col.replace(' ', '_')
            new_col = new_col.upper()
            new_cols.append(new_col)
        else:
            new_cols.append(col)
    df.columns = new_cols

    return df

def formatting(df: pd.DataFrame, mapping: list[dict]) -> pd.DataFrame:
    for m in mapping:
        cols = m.get('Column', [])
        trans_type = m.get('trans_type', None)

        if trans_type == 'str':
            for col in cols:
                if col in df.columns:
                    df[col] = df[col].astype(str)
                    df[col] = df[col].str.upper()
                    df[col] = df[col].str.strip()
                    df[col] = df[col].str.replace(" ", "_")

        elif trans_type == 'int':
            for col in cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

        elif trans_type == 'float':
            for col in cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").round(4)

        elif trans_type == 'bool':
            for col in cols:
                if col in df.columns:
                    df[col] = df[col].astype("boolean")

        elif trans_type == 'percent':
            for col in cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce") * 100
                    df[col] = df[col].round(2)
                    df[col] = df[col].apply(lambda x: f"{x}%" if pd.notna(x) else pd.NA)

        elif trans_type == 'date':
            date_format = m.get("format", "YYYY-MM-DD")

            from dateutil import parser
            def parse_date(val):
                try:
                    return parser.parse(str(val), dayfirst=False, yearfirst=False)
                except Exception:
                    return pd.NaT

            for col in cols:
                if col in df.columns:
                    df[col] = df[col].apply(parse_date)

                    if date_format == "YYYY-MM-DD":
                        df[col] = df[col].dt.strftime("%Y-%m-%d")
                    elif date_format == "DD_MM_YY":
                        df[col] = df[col].dt.strftime("%d-%m-%y")
                    elif date_format == "MM-YY":
                        df[col] = df[col].dt.strftime("%m-%y")

        elif trans_type == "scale":
            factor = m.get("factor", 1)
            operation = m.get("operation", "mul")
            for col in cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce")  # 转数值
                    if operation == "mul":
                        df[col] = df[col] * factor
                    elif operation == "div":
                        df[col] = df[col] / factor
                    elif operation == "add":
                        df[col] = df[col] + factor
                    elif operation == "sub":
                        df[col] = df[col] - factor


        elif trans_type == 'missing':
            strategy = m.get("strategy", None)
            for col in cols:
                if col in df.columns:
                    if strategy == "mean":
                        df[col] = df[col].fillna(df[col].mean())
                    elif strategy == "median":
                        df[col] = df[col].fillna(df[col].median())
                    elif strategy == "nan":
                        df[col] = df[col].fillna(pd.NA)

        elif trans_type == "outlier":
            method = m.get("method", "zscore")
            replace = m.get("replace", "nan")
            threshold = m.get("threshold", 3)
            for col in cols:
                if col in df.columns:
                    series = df[col]
                    if method == "zscore":
                        mean, std = series.mean(), series.std()
                        mask = abs(series - mean) > threshold * std
                    elif method == "iqr":
                        q1, q3 = series.quantile([0.25, 0.75])
                        iqr = q3 - q1
                        mask = (series < q1 - threshold * iqr) | (series > q3 + threshold * iqr)

                    if replace == "mean":
                        df.loc[mask, col] = mean
                    elif replace == "median":
                        df.loc[mask, col] = series.median()
                    elif replace == "nan":
                        df.loc[mask, col] = pd.NA

    df = df.astype(object).where(df.notna(), float("nan"))
    return df

def diff_highlight(df1: pd.DataFrame, df2: pd.DataFrame, mapping: list[dict]) -> None:
    for m in mapping:
        suffix1 = m.get('suffix1', [])
        suffix2 = m.get('suffix2', [])
        out_file = m.get('out_file', [])

    df1_renamed = df1.add_suffix(f"_{suffix1}") if suffix1 else df1.copy()
    df2_renamed = df2.add_suffix(f"_{suffix2}") if suffix2 else df2.copy()

    merged = pd.DataFrame()
    for col in df1.columns:
        merged[f"{col}_{suffix1}" if suffix1 else col] = df1[col]
        merged[f"{col}_{suffix2}" if suffix2 else col] = df2[col]
    merged.to_excel(out_file, index=False)

    wb = load_workbook(out_file)
    ws = wb.active

    fill_up = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_down = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for j, col in enumerate(df1.columns):
        col1 = 2 * j + 1
        col2 = 2 * j + 2

        for i in range(len(df1)):
            val1 = df1.iloc[i, j]
            val2 = df2.iloc[i, j]

            if pd.notna(val1) and pd.notna(val2) and val1 != val2:
                cell = ws.cell(row=i + 2, column=col2)
                if val2 > val1:
                    cell.fill = fill_up
                else:
                    cell.fill = fill_down

    wb.save(out_file)

def write_diff_report(df1: pd.DataFrame, df2: pd.DataFrame, mapping: list[dict]) -> None:
    for m in mapping:
        sheet_suffix1 = m.get('sheet_suffix1', '_1')
        sheet_suffix2 = m.get('sheet_suffix2', '_2')
        out_file = m.get('out_file', '')

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name=f"DataFrame_{sheet_suffix1}", index=False)
        df2.to_excel(writer, sheet_name=f"DataFrame_{sheet_suffix2}", index=False)

    wb = load_workbook(out_file)
    ws2 = wb[f"DataFrame_{sheet_suffix2}"]

    fill_up = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_down = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    diffs = []
    common_cols = [col for col in df1.columns if col in df2.columns]

    for col in common_cols:
        col_idx = df2.columns.get_loc(col) + 1
        for r in range(len(df1)):
            val1 = df1.iloc[r][col]
            val2 = df2.iloc[r][col]

            if pd.notna(val1) and pd.notna(val2) and val1 != val2:
                row_idx = r + 2
                if val2 > val1:
                    ws2.cell(row=row_idx, column=col_idx).fill = fill_up
                    change = "Up"
                else:
                    ws2.cell(row=row_idx, column=col_idx).fill = fill_down
                    change = "Down"

                diffs.append({
                    "Row": row_idx,
                    "Column": col,
                    f"Old{sheet_suffix1}": val1,
                    f"New{sheet_suffix2}": val2,
                    "Change": change
                })

    wb.save(out_file)

    if diffs:
        df_diff = pd.DataFrame(diffs)
        with pd.ExcelWriter(out_file, engine="openpyxl", mode="a") as writer:
            df_diff.to_excel(writer, sheet_name="Diff Summary", index=False)

def combine_df(df1: pd.DataFrame, df2: pd.DataFrame, mapping: list[dict]) -> pd.DataFrame:
    '''
    mapping = [{"method": "merge", "on": ["ID"]}]
    or
    mapping = [{"method": "concat"}]
    '''
    method = mapping[0]["method"].lower()

    if method == "concat":
        a, b = df1.copy(), df2.copy()
        a.columns = [str(c).strip().upper() for c in a.columns]
        b.columns = [str(c).strip().upper() for c in b.columns]

        return pd.concat([a, b], axis = 0, ignore_index = True, sort = False)

    elif method == "merge":

        a, b = df1.copy(), df2.copy()
        if "on" not in mapping[0]:
            raise ValueError("merge须指定公共参数，例如：[{'method': 'merge', 'on': ['ID']}]")

        on_cols = mapping[0]["on"]

        for col in on_cols:
            if col not in a.columns:
                a[col] = None
            if col not in b.columns:
                b[col] = None

        for c in on_cols:
            a[c] = a[c].astype(str).str.strip()
            b[c] = b[c].astype(str).str.strip()

        return a.merge(b, on = on_cols, suffixes = ("_L", "_R"))

def export_data(df: pd.DataFrame, filename: str = 'output.csv', file_format: str = None) -> None:
    """
    通用导出函数：支持 CSV, Excel, JSON
    :param df: 需要导出的 DataFrame
    :param filename: 文件名，默认 output.csv
    :param file_format: 文件格式，可选 'csv', 'excel', 'json'，默认根据文件扩展名判断
    """
    # 如果没有指定格式，则从文件名扩展名推断
    if file_format is None:
        if filename.lower().endswith('.csv'):
            file_format = 'csv'
        elif filename.lower().endswith(('.xls', '.xlsx')):
            file_format = 'excel'
        elif filename.lower().endswith('.json'):
            file_format = 'json'
        else:
            raise ValueError("无法识别文件格式，请指定 file_format 参数 ('csv', 'excel', 'json')")

    # 根据不同格式导出
    if file_format == 'csv':
        df.to_csv(filename, index=False, encoding='utf-8')
    elif file_format == 'excel':
        df.to_excel(filename, index=False, engine='openpyxl')
    elif file_format == 'json':
        df.to_json(filename, orient='records', force_ascii=False)
    else:
        raise ValueError("不支持的文件格式: {}".format(file_format))
